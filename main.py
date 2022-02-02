import tkinter as tk
from tkinter import ttk
from datetime import datetime
from openpyxl import Workbook, load_workbook
from string import ascii_lowercase as al
from tkinter import messagebox

try:
    wb = load_workbook("interest.xlsx")
    ws = wb.active
except FileNotFoundError:
    wb = Workbook()
    ws = wb.active
    titles = [
        "Sr.No",
        "Sale date",
        "Amt",
        "Received date",
        "Total days",
        "Grace days",
        "Chargeable days",
        "Rate of interest",
        "Interest amt",
        "Grand total"
    ]
    for n in range(len(titles)):
        ws[f"{al[n]}1"] = titles[n]

window = tk.Tk()
window.geometry("400x500")
window.counter = ws.max_row

saleDate_var = tk.StringVar()
amt_var = tk.StringVar()
receivedDate_var = tk.StringVar()
graceDays_var = tk.StringVar()
interest_var = tk.StringVar(window, "1")


def get_vals():
    if check_all_vals():
        saleDate = datetime.strptime(saleDate_var.get(), "%d/%m/%y")
        receivedDate = datetime.strptime(receivedDate_var.get(), "%d/%m/%y")
        graceDays = float(graceDays_var.get())
        amt = float(amt_var.get())
        interest = float(interest_var.get())

        if interest < 3:
            interest_amt = ((amt * interest / 100) / 30) * ((receivedDate - saleDate).days - graceDays)
        else:
            interest_amt = ((amt * interest / 100) / 365) * ((receivedDate - saleDate).days - graceDays)

        ws.append(
            [int(window.counter), saleDate.date(), float(amt), receivedDate.date(), int((receivedDate - saleDate).days),
             int(graceDays), int((receivedDate - saleDate).days - graceDays), float(interest), float(interest_amt),
             float(amt + interest_amt)])

        try:
            saleDate_var.set('')
            amt_var.set('')
            receivedDate_var.set('')
            graceDays_var.set('')
            interest_var.set('1')
            window.counter += 1
            Sno_label.configure(text=f"S.No {window.counter}")
            interest_amt_label.configure(text=f"Interest amount = {round(interest_amt, 2)}")
            total_amt_label.configure(text=f"Total amount = {amt + round(interest_amt, 2)}")
            wb.save("interest.xlsx")
        except PermissionError as e:
            messagebox.showerror(str(e), "Close the file before submitting")


def check_vals(var):
    if "Date" in var:
        try:
            datetime.strptime(eval(f"{var}.get()"), "%d/%m/%y")
            return True
        except ValueError:
            return False
    else:
        try:
            float(eval(f"{var}.get()"))
            return True
        except ValueError:
            return False


def check_all_vals():
    try:
        datetime.strptime(saleDate_var.get(), "%d/%m/%y")
        datetime.strptime(receivedDate_var.get(), "%d/%m/%y")
        float(graceDays_var.get())
        float(amt_var.get())
        return True
    except ValueError:
        return False


def Focus(key, check):
    eval(f"{key}.focus()")
    if not check_vals(check):
        eval(f"{check.split('_')[0]}_error.configure(text='Error', foreground='red')")
    else:
        eval(f"{check.split('_')[0]}_error.configure(text='', foreground='red')")


Sno_label = ttk.Label(window, text=f"S.No {window.counter}")
Sno_label.grid(row=1, column=0)

saleDate_label = ttk.Label(window, text="Sale Date")
saleDate_label.grid(row=2, column=0)
saleDate_entry = ttk.Entry(window, textvariable=saleDate_var)
saleDate_entry.grid(row=2, column=1)

amt_label = ttk.Label(window, text="Amount")
amt_label.grid(row=3, column=0)
amt_entry = ttk.Entry(window, textvariable=amt_var)
amt_entry.grid(row=3, column=1)

receivedDate_label = ttk.Label(window, text="Received Date")
receivedDate_label.grid(row=4, column=0)
receivedDate_entry = ttk.Entry(window, textvariable=receivedDate_var)
receivedDate_entry.grid(row=4, column=1)

graceDays_label = ttk.Label(window, text="Grace Days")
graceDays_label.grid(row=5, column=0)
graceDays_entry = ttk.Entry(window, textvariable=graceDays_var)
graceDays_entry.grid(row=5, column=1)

interest_label = ttk.Label(window, text="Interest")
interest_label.grid(row=6, column=0)
rb1 = ttk.Radiobutton(window, text="1.5%", variable=interest_var, value="1.5")
rb2 = ttk.Radiobutton(window, text="2%   ", variable=interest_var, value="2")
rb3 = ttk.Radiobutton(window, text="18% ", variable=interest_var, value="18")
rb4 = ttk.Radiobutton(window, text="24% ", variable=interest_var, value="24")
rb1.grid(row=6, column=1)
rb2.grid(row=7, column=1)
rb3.grid(row=8, column=1)
rb4.grid(row=9, column=1)

interest_amt_label = ttk.Label(window, text="")
interest_amt_label.grid(row=10, column=0)

total_amt_label = ttk.Label(window, text="")
total_amt_label.grid(row=11, column=0)

submitButton = ttk.Button(window, text="Submit", command=get_vals)
submitButton.grid(row=12, column=2)

# Error labels
saleDate_error = ttk.Label(window, text="")
saleDate_error.grid(row=2, column=2)

amt_error = ttk.Label(window, text="")
amt_error.grid(row=3, column=2)

receivedDate_error = ttk.Label(window, text="")
receivedDate_error.grid(row=4, column=2)

graceDays_error = ttk.Label(window, text="")
graceDays_error.grid(row=5, column=2)

# Binds
saleDate_entry.bind("<Up>", lambda i: Focus("submitButton", "saleDate_entry"))
saleDate_entry.bind("<Down>", lambda i: Focus("amt_entry", "saleDate_entry"))

amt_entry.bind("<Up>", lambda i: Focus("saleDate_entry", "amt_entry"))
amt_entry.bind("<Down>", lambda i: Focus("receivedDate_entry", "amt_entry"))

receivedDate_entry.bind("<Up>", lambda i: Focus("amt_entry", "receivedDate_entry"))
receivedDate_entry.bind("<Down>", lambda i: Focus("graceDays_entry", "receivedDate_entry"))

graceDays_entry.bind("<Up>", lambda i: Focus("receivedDate_entry", "graceDays_entry"))
graceDays_entry.bind("<Down>", lambda i: Focus("rb1", "graceDays_entry"))

submitButton.bind("<Up>", lambda i: rb4.focus())
submitButton.bind("<Down>", lambda i: saleDate_entry.focus())

window.mainloop()
